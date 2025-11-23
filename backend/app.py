"""
Store343 Backend API
Flask server for processing documents with Claude AI
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
import anthropic
import os
import base64
import json
import traceback
from typing import List, Dict, Any
from openpyxl import load_workbook
from io import BytesIO
import fitz  # PyMuPDF

app = Flask(__name__)
CORS(app)

# Claude API client - strip whitespace from API key to handle copy-paste errors
api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
client = anthropic.Anthropic(api_key=api_key)

def excel_to_text(base64_data: str) -> str:
    """
    Convert Excel file (base64) to formatted text for Claude API

    Args:
        base64_data: Base64 encoded Excel file

    Returns:
        Formatted text representation of the Excel data
    """
    try:
        # Decode base64
        excel_bytes = base64.b64decode(base64_data)

        # Load workbook
        wb = load_workbook(filename=BytesIO(excel_bytes), data_only=True)

        # Get first sheet (or active sheet)
        ws = wb.active

        # Convert to text format
        text_lines = []
        text_lines.append(f"Excel Document: {wb.sheetnames[0] if wb.sheetnames else 'Sheet1'}")
        text_lines.append("=" * 80)
        text_lines.append("")

        # Process rows
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # Skip completely empty rows
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            # Format row data
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append("")
                else:
                    row_data.append(str(cell).strip())

            # Join cells with | separator
            text_lines.append(" | ".join(row_data))

        text_lines.append("")
        text_lines.append("=" * 80)
        text_lines.append(f"Total rows: {ws.max_row}")

        return "\n".join(text_lines)

    except Exception as e:
        raise Exception(f"Failed to parse Excel file: {str(e)}")

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({"status": "healthy", "service": "Store343 API"}), 200

@app.route('/api/process-napi-info', methods=['POST'])
def process_napi_info():
    """
    Process Napi Info document (image or PDF) with Claude AI

    Request body:
    {
        "image_base64": "base64_encoded_image_data",
        "image_type": "image/jpeg" or "application/pdf"
    }

    Response:
    {
        "success": true,
        "blocks": [
            {
                "tema": "Topic",
                "erintett": "Affected area",
                "tartalom": "Content",
                "hatarido": "Deadline",
                "emoji": "📋",
                "checkboxes": ["Task 1", "Task 2"],
                "images": []
            }
        ],
        "usage": {
            "input_tokens": 1234,
            "output_tokens": 567
        }
    }
    """
    try:
        print("📄 [NAPI] process-napi-info endpoint called")
        data = request.get_json()
        print(f"📄 [NAPI] Request data keys: {list(data.keys()) if data else 'None'}")

        # Accept both 'image_base64' (images) and 'document_base64' (PDFs)
        image_base64 = data.get('image_base64') or data.get('document_base64')
        if not image_base64:
            print("❌ [NAPI] Missing image_base64 or document_base64")
            return jsonify({
                "success": False,
                "error": "Missing image_base64 or document_base64 in request"
            }), 400

        # Accept both 'image_type' and 'document_type'
        image_type = data.get('image_type') or data.get('document_type', 'image/jpeg')
        print(f"📄 [NAPI] Document type: {image_type}, base64 length: {len(image_base64)}")

        # Convert PDF to PNG if needed (Claude API only accepts image formats)
        pdf_pages = []  # Will store base64 PNG images for each page
        if image_type == 'application/pdf':
            print("📄 [NAPI] Converting PDF to PNG...")
            try:
                # Decode base64 PDF
                pdf_bytes = base64.b64decode(image_base64)

                # Open PDF with PyMuPDF
                pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")
                print(f"📄 [NAPI] PDF has {len(pdf_document)} pages")

                # Convert ALL pages to PNG
                for page_num in range(len(pdf_document)):
                    page = pdf_document[page_num]

                    # Render page to pixmap (image) at 2x resolution for better quality
                    mat = fitz.Matrix(2.0, 2.0)  # 2x zoom for better OCR
                    pix = page.get_pixmap(matrix=mat)

                    # Convert pixmap to PNG bytes
                    png_bytes = pix.tobytes("png")

                    # Encode to base64 and store
                    page_base64 = base64.b64encode(png_bytes).decode('utf-8')
                    pdf_pages.append(page_base64)

                    print(f"📄 [NAPI] Page {page_num + 1} converted to PNG, base64 length: {len(page_base64)}")

                pdf_document.close()
                image_type = 'image/png'
                print(f"📄 [NAPI] All {len(pdf_pages)} pages converted to PNG")
            except Exception as e:
                print(f"❌ [NAPI] PDF conversion error: {str(e)}")
                traceback.print_exc()
                return jsonify({
                    "success": False,
                    "error": f"Failed to convert PDF: {str(e)}"
                }), 500
        else:
            # For regular images, just use the single image
            pdf_pages = [image_base64]

        # Prepare messages for Claude
        prompt = """Analyze this Hungarian LIDL Napi Információ PDF document. Extract ALL topics/sections.

DOCUMENT STRUCTURE:
- Header: "Napi Információ" + date (e.g., "2025. november 20., csütörtök")
- Multiple topics, each with:
  ☑ Checkboxes at top (Info, Feladat, Mindenki, Jelentés, Melléklet)
  Téma: [topic title]
  Érintett: [affected people/department]
  [Content - can include text, tables, product lists]
  Határidő: [deadline or missing]

FOR EACH TOPIC, EXTRACT:
{
  "tema": "exact topic title from 'Téma:' line",
  "erintett": "exact text from 'Érintett:' line",
  "tartalom": "full content - preserve lists with bullet points (•), tables with structure",
  "hatarido": "YYYY-MM-DD HH:MM or null",
  "emoji": "relevant emoji (🛒📦💰🍺📊🗂️📋📝🧾)",
  "checkboxes": ["Info", "Feladat", "Mindenki", "Jelentés", "Melléklet"],
  "images": []
}

CHECKPOINT EXTRACTION RULES:
- Look for ☑ checkmarks at the START of each topic
- Common patterns: "☑ Info ☑ Feladat", "☐ Info ☑ Feladat ☑ Mindenki"
- Only include checkboxes that have ☑ (checked) mark
- Empty checkbox ☐ = not included

DEADLINE NORMALIZATION:
- "ma este zárás" → today at 21:00
- "hétfő este zárás volt!" → last Monday at 21:00
- "2025.11.20. (nyitás)" → "2025-11-20 06:00"
- "2025.11.22. (szombat)" → "2025-11-22 00:00"
- "vásárnapig" → next Sunday at 00:00
- If no deadline mentioned → null

CONTENT FORMATTING:
- Product lists: preserve with bullets "• 4893 Házánk Kincsei Téli szalámi..."
- Tables: format clearly with line breaks between rows
- Bizonylat numbers: keep highlighted "86888 visszaküldési bizonylatszámon"
- Preserve ALL Hungarian text exactly as written

EMOJI SELECTION:
- "Mopro akciós sarok" → 🛒
- "visszaküldés" or "szortiment" → 📦
- "Lidl Plus" or "termékek" or "árak" → 💰
- "sör" → 🍺
- "forgalás" → 📊
- "készletjelentés" or "MOHU" → 🗂️
- "munkaterv" or "BV" → 📋
- "bejárás" or "jegyzőkönyv" → 📝
- "készletszámolás" → 🧾

Return ONLY valid JSON array, no markdown, no explanation:
[{"tema": "...", "erintett": "...", "tartalom": "...", "hatarido": "...", "emoji": "...", "checkboxes": [...], "images": []}]

CRITICAL: Extract ALL topics from ALL pages of the PDF!"""

        # Build content array with all pages
        content_items = []
        for page_num, page_base64 in enumerate(pdf_pages):
            print(f"📄 [NAPI] Adding page {page_num + 1} to Claude API request")
            content_items.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": image_type,
                    "data": page_base64,
                },
            })

        # Add prompt at the end
        content_items.append({
            "type": "text",
            "text": prompt
        })

        print(f"📄 [NAPI] Calling Claude API with {len(pdf_pages)} page(s)...")
        message = client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=16384,  # Claude Sonnet 4.5 supports up to 64K output tokens
            messages=[
                {
                    "role": "user",
                    "content": content_items,
                }
            ],
        )
        print(f"📄 [NAPI] Claude API response received. Tokens: input={message.usage.input_tokens}, output={message.usage.output_tokens}")

        # Extract JSON from response
        response_text = message.content[0].text.strip()

        # Remove markdown code blocks if present
        if response_text.startswith('```'):
            response_text = response_text.split('```')[1]
            if response_text.startswith('json'):
                response_text = response_text[4:]
            response_text = response_text.strip()

        print(f"📄 [NAPI] Parsing JSON response (length: {len(response_text)})")
        blocks = json.loads(response_text)
        print(f"📄 [NAPI] Successfully parsed {len(blocks)} blocks")

        return jsonify({
            "success": True,
            "blocks": blocks,
            "usage": {
                "input_tokens": message.usage.input_tokens,
                "output_tokens": message.usage.output_tokens
            }
        }), 200

    except json.JSONDecodeError as e:
        print(f"❌ [NAPI] JSON decode error: {str(e)}")
        print(f"❌ [NAPI] Response text: {response_text[:500]}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"Failed to parse AI response as JSON: {str(e)}"
        }), 500
    except Exception as e:
        print(f"❌ [NAPI] Exception: {type(e).__name__}: {str(e)}")
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"{type(e).__name__}: {str(e)}"
        }), 500

@app.route('/api/process-nf-visszakuldes', methods=['POST'])
def process_nf_visszakuldes():
    """
    Process NF visszaküldés document (Excel/image/PDF) with Claude AI

    Request body:
    {
        "document_base64": "base64_encoded_document_data",
        "document_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or "image/jpeg" or "application/pdf"
    }

    Response:
    {
        "success": true,
        "termekek": [
            {
                "cikkszam": "123456",
                "cikk_megnevezes": "Product name",
                "bizonylat_szam": "33606",
                "elvi_keszlet": 5
            }
        ],
        "usage": {
            "input_tokens": 1234,
            "output_tokens": 567
        }
    }
    """
    try:
        print("🔵 [NF] Received request to /api/process-nf-visszakuldes")
        data = request.get_json()
        print(f"🔵 [NF] Request data keys: {data.keys() if data else 'None'}")

        if not data or 'document_base64' not in data:
            print("❌ [NF] Missing document_base64 in request")
            return jsonify({
                "success": False,
                "error": "Missing document_base64 in request"
            }), 400

        document_base64 = data['document_base64']
        document_type = data.get('document_type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        print(f"🔵 [NF] Document type: {document_type}")
        print(f"🔵 [NF] Document base64 length: {len(document_base64)} chars")

        # Prepare messages for Claude
        prompt = """Analyze this Hungarian LIDL NF visszaküldés (return) document. Extract all product information.

The document contains products with these columns:
- Cikkszám (product code/SKU)
- Cikk megnevezés (product name)
- Bizonylat (receipt/document number)
- Mennyiség (quantity - this is "elvi készlet" / theoretical stock)

Some rows are category headers (like "Parkside", "PLU") - IGNORE these rows (they don't have a cikkszám).

Return ONLY valid JSON array of products, nothing else:
[{"cikkszam": "...", "cikk_megnevezes": "...", "bizonylat_szam": "...", "elvi_keszlet": 0}]

Important:
- Extract ALL products from ALL bizonylatok
- Skip category header rows
- If mennyiség is empty, use 0
- Use exact Hungarian text
- cikkszam should be a string (may have leading zeros)
- bizonylat_szam should be a string
- elvi_keszlet should be an integer (0 if empty/null)"""

        # Determine content type and prepare for Claude API
        content_items = []

        if document_type.startswith('image/'):
            # Image type - use image content
            print("🔵 [NF] Using 'image' content type for Claude API")
            content_items.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": document_type,
                    "data": document_base64,
                },
            })
        elif document_type == 'application/pdf':
            # PDF type - use document content
            print("🔵 [NF] Using 'document' content type for Claude API (PDF)")
            content_items.append({
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": document_type,
                    "data": document_base64,
                },
            })
        elif 'spreadsheet' in document_type or 'excel' in document_type:
            # Excel type - convert to text first
            print("🔵 [NF] Converting Excel to text for Claude API")
            try:
                excel_text = excel_to_text(document_base64)
                print(f"🔵 [NF] Excel converted to text: {len(excel_text)} chars")
                print(f"🔵 [NF] Excel preview:\n{excel_text[:500]}...")

                # Add as text content with the Excel data
                content_items.append({
                    "type": "text",
                    "text": f"Here is the Excel spreadsheet data:\n\n{excel_text}\n\n{prompt}"
                })
            except Exception as e:
                print(f"❌ [NF] Excel conversion failed: {str(e)}")
                raise Exception(f"Failed to parse Excel file: {str(e)}")
        else:
            # Other document types - try as document
            print(f"🔵 [NF] Using 'document' content type for {document_type}")
            content_items.append({
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": document_type,
                    "data": document_base64,
                },
            })

        # Add prompt as text (unless already added with Excel data)
        if len(content_items) > 0 and content_items[0].get("type") != "text":
            content_items.append({
                "type": "text",
                "text": prompt
            })

        print("🔵 [NF] Calling Claude API...")
        message = client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=16384,  # Claude Sonnet 4.5 limit (64K available, using 16K for efficiency)
            messages=[
                {
                    "role": "user",
                    "content": content_items,
                }
            ],
        )

        print("✅ [NF] Claude API call successful!")
        print(f"🔵 [NF] Token usage - Input: {message.usage.input_tokens}, Output: {message.usage.output_tokens}")

        # Extract JSON from response
        response_text = message.content[0].text.strip()
        print(f"🔵 [NF] Raw response length: {len(response_text)} chars")
        print(f"🔵 [NF] Response preview: {response_text[:200]}...")

        # Try to extract JSON from response (Claude sometimes adds explanatory text)
        json_text = response_text

        # Remove markdown code blocks if present
        if '```' in json_text:
            print("🔵 [NF] Removing markdown code blocks from response")
            parts = json_text.split('```')
            if len(parts) >= 2:
                json_text = parts[1]
                if json_text.startswith('json'):
                    json_text = json_text[4:]
                json_text = json_text.strip()

        # If response starts with text, try to find the JSON array
        if not json_text.startswith('['):
            print("🔵 [NF] Response has prefix text, extracting JSON array")
            # Find the first [ and take everything from there
            bracket_index = json_text.find('[')
            if bracket_index != -1:
                json_text = json_text[bracket_index:]
                print(f"🔵 [NF] Extracted JSON starting at position {bracket_index}")

        # Check if response was truncated (hit max_tokens limit)
        if message.usage.output_tokens >= 16300:  # Close to max_tokens (16384)
            print("⚠️ [NF] Response may be truncated (hit max_tokens limit)")

        # Try to fix incomplete JSON if response was truncated
        if not json_text.endswith(']'):
            print("⚠️ [NF] Response doesn't end with ], attempting to fix")

            # Find the last complete object
            # Look for the last complete "}" that closes an object
            last_complete_obj = json_text.rfind('},')
            if last_complete_obj != -1:
                # Cut after the complete object and close the array
                json_text = json_text[:last_complete_obj + 1] + '\n]'
                print(f"🔵 [NF] Fixed truncated JSON at position {last_complete_obj}")
            else:
                # Try to find at least one complete object
                last_brace = json_text.rfind('}')
                if last_brace != -1:
                    json_text = json_text[:last_brace + 1] + '\n]'
                    print(f"🔵 [NF] Fixed truncated JSON at last brace {last_brace}")

        print("🔵 [NF] Parsing JSON response...")
        print(f"🔵 [NF] JSON text preview: {json_text[:200]}...")
        print(f"🔵 [NF] JSON text suffix: ...{json_text[-100:]}")
        termekek = json.loads(json_text)
        print(f"✅ [NF] Successfully parsed {len(termekek)} termekek")

        return jsonify({
            "success": True,
            "termekek": termekek,
            "usage": {
                "input_tokens": message.usage.input_tokens,
                "output_tokens": message.usage.output_tokens
            }
        }), 200

    except json.JSONDecodeError as e:
        print(f"❌ [NF] JSON decode error: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"Failed to parse AI response as JSON: {str(e)}"
        }), 500
    except Exception as e:
        print(f"❌ [NF] Exception: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"{type(e).__name__}: {str(e)}"
        }), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
